# -*- coding: utf-8 -*-
"""Build an alumni import sheet from the two roster spreadsheets.

    python scripts/roster/extract_roster.py

Reads the .xlsx files in members_do_not_commit/ and writes, next to them:

    import_review.csv   one row per importable person, for you to correct
    skipped.csv         people who cannot be imported, with the reason

Nothing is sent anywhere -- this only reads and writes local files. Review
import_review.csv (especially every row where needs_review is yes), then feed
it to scripts/roster/import_alumni.mjs.

Requires: pip install openpyxl
"""
import csv
import datetime
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is required: pip install openpyxl')

ROSTER_DIR = 'members_do_not_commit'
FILE_A = 'נתוני לפידים עבור כנס מחזור קיץ 2026.xlsx'
FILE_B = 'שמות בוגרי לפידים + פעילים בתכנית 28.1.26 עבור מיפוי לינקדאין ומיילים.xlsx'
SHEET_MAIN = 'גיליון1'
SHEET_NO_EMAIL = 'כתובות מייל חסרות'

# File A column indexes (0-based), header on row 2
COL_COHORT, COL_NAME, COL_EMAIL_1, COL_EMAIL_2 = 1, 2, 12, 13
# File B
COL_B_NAME, COL_B_LINKEDIN = 2, 3

EMAIL_RE = re.compile(r'[\w.+-]+@[\w-]+(?:\.[\w-]+)+')
YEAR_RE = re.compile(r'(?:19|20)\d{2}')
HEADER_NAMES = {'שם', 'שם מלא'}

# Campus addresses stop working after graduation, so a personal address wins
# when someone lists both.
EXPIRING_DOMAINS = ('campus.technion.ac.il', 'technion.ac.il', 'alumni.technion.ac.il')


def clean(value):
    if value is None:
        return None
    text = str(value).replace('\xa0', ' ').strip()
    return text or None


def norm_name(value):
    text = clean(value)
    return re.sub(r'\s+', ' ', text) if text else None


def emails_in(*cells):
    """Pull every address out of cells that may hold two, separated by ; or |."""
    found = []
    for cell in cells:
        if cell:
            found += [m.group(0).lower() for m in EMAIL_RE.finditer(str(cell))]
    ordered = []
    for address in found:
        if address not in ordered:
            ordered.append(address)
    return ordered


def pick_login_email(addresses):
    """Prefer an address that will still work in five years."""
    for address in addresses:
        if not address.endswith(EXPIRING_DOMAINS):
            return address
    return addresses[0]


# --- Hebrew -> Latin ---------------------------------------------------------
# Hebrew names are written without vowels, so this can only ever be a starting
# point: שני comes out "Shni", not "Shani". Every transliterated row is flagged
# for review.
FINAL_TO_BASE = {'ך': 'כ', 'ם': 'מ', 'ן': 'נ', 'ף': 'פ', 'ץ': 'צ'}
LETTERS = {
    'א': 'a', 'ב': 'b', 'ג': 'g', 'ד': 'd', 'ה': 'h', 'ו': 'o', 'ז': 'z',
    'ח': 'ch', 'ט': 't', 'י': 'i', 'כ': 'k', 'ל': 'l', 'מ': 'm', 'נ': 'n',
    'ס': 's', 'ע': 'a', 'פ': 'p', 'צ': 'tz', 'ק': 'k', 'ר': 'r', 'ש': 'sh',
    'ת': 't',
}


def transliterate_word(word):
    chars = [FINAL_TO_BASE.get(c, c) for c in word]
    out = []
    for i, c in enumerate(chars):
        first, last = i == 0, i == len(chars) - 1
        if c == 'ו':
            # וו is a consonant v; a lone ו is usually the vowel o/u
            if (i + 1 < len(chars) and chars[i + 1] == 'ו') or (i > 0 and chars[i - 1] == 'ו'):
                out.append('v' if first or (i > 0 and chars[i - 1] != 'ו') else '')
            else:
                out.append('v' if first else 'o')
        elif c == 'י':
            out.append('y' if first else 'i')
        elif c == 'ה':
            out.append('a' if last and not first else 'h')  # final ה marks a vowel
        elif c == 'א':
            out.append('' if last and not first else 'a')
        elif c == 'ע':
            out.append('' if last else 'a')
        elif c in LETTERS:
            out.append(LETTERS[c])
        elif c in ('"', "'", '`', '״', '׳'):
            continue
        else:
            out.append(c)
    word = ''.join(out)
    word = re.sub(r'([aeiou])\1+', r'\1', word)  # collapse doubled vowels
    return word.capitalize()


def transliterate(hebrew):
    parts = re.split(r'([ \-])', hebrew)
    return ''.join(transliterate_word(p) if p.strip() and p not in ' -' else p
                   for p in parts).strip()


def name_from_slug(url):
    """'ari-zigler-bab15655' -> 'Ari Zigler'. Returns None if it will not split."""
    slug = url.rstrip('/').split('/')[-1].split('?')[0]
    slug = re.sub(r'-[0-9a-z]*\d[0-9a-z]*$', '', slug)  # trailing uniquifier
    parts = [p for p in re.split(r'[-_.]', slug) if p and not p.isdigit()]
    if len(parts) < 2:
        return None
    if any(len(p) < 2 for p in parts):
        return None
    return ' '.join(p.capitalize() for p in parts)


# --- read the spreadsheets ---------------------------------------------------
def read_file_a(path):
    """Three stacked tables share this sheet; the last repeats earlier people,
    so rows are merged by email and duplicates collapse on their own."""
    sheet = openpyxl.load_workbook(path, data_only=True)[SHEET_MAIN]
    people, no_email, cohort = {}, [], None

    for row in sheet.iter_rows(min_row=3, values_only=True):
        cohort_cell = clean(row[COL_COHORT])
        if cohort_cell:
            match = YEAR_RE.search(cohort_cell)
            if match:
                cohort = int(match.group(0))
        name = norm_name(row[COL_NAME])
        if not name or name in HEADER_NAMES:
            continue

        addresses = emails_in(row[COL_EMAIL_1], row[COL_EMAIL_2])
        if not addresses:
            no_email.append({'name': name, 'cohort': cohort, 'reason': 'no email in file A'})
            continue

        key = pick_login_email(addresses)
        person = people.setdefault(key, {'name': name, 'cohort': cohort, 'emails': []})
        person['cohort'] = person['cohort'] or cohort
        for address in addresses:
            if address not in person['emails']:
                person['emails'].append(address)
    return people, no_email


def read_file_b(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    links = {}
    for row in workbook[SHEET_MAIN].iter_rows(min_row=3, values_only=True):
        name, url = norm_name(row[COL_B_NAME]), clean(row[COL_B_LINKEDIN])
        if name and url and 'linkedin.com' in url:
            links[name] = url
    missing = []
    if SHEET_NO_EMAIL in workbook.sheetnames:
        missing = [norm_name(r[0]) for r in workbook[SHEET_NO_EMAIL].iter_rows(values_only=True)
                   if norm_name(r[0])]
    return links, missing


def main():
    path_a = os.path.join(ROSTER_DIR, FILE_A)
    path_b = os.path.join(ROSTER_DIR, FILE_B)
    for path in (path_a, path_b):
        if not os.path.exists(path):
            sys.exit('missing spreadsheet: %s' % path)

    people, no_email = read_file_a(path_a)
    links, missing_sheet = read_file_b(path_b)

    rows = []
    for login_email, person in people.items():
        linkedin = links.get(person['name'])
        english = name_from_slug(linkedin) if linkedin else None
        source = 'linkedin'
        if not english:
            english = transliterate(person['name'])
            source = 'transliterated'
        # A slug with fewer parts than the Hebrew name has dropped something --
        # usually a middle name -- so it still wants a human look.
        parts_differ = len(english.split()) != len(person['name'].split())
        rows.append({
            'email': login_email,
            'full_name_he': person['name'],
            'full_name_en': english,
            'source': source,
            'needs_review': 'no' if source == 'linkedin' and not parts_differ else 'yes',
            'cohort_start': person['cohort'] or '',
            'linkedin_url': linkedin or '',
            'other_emails': ' '.join(e for e in person['emails'] if e != login_email),
        })
    rows.sort(key=lambda r: (str(r['cohort_start']), r['full_name_he']))

    os.makedirs(ROSTER_DIR, exist_ok=True)
    review_path = os.path.join(ROSTER_DIR, 'import_review.csv')
    fields = ['email', 'full_name_he', 'full_name_en', 'source', 'needs_review',
              'cohort_start', 'linkedin_url', 'other_emails']
    # utf-8-sig so Excel shows the Hebrew correctly instead of mojibake
    with open(review_path, 'w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    skipped = [dict(r, reason=r['reason']) for r in no_email]
    known = {r['full_name_he'] for r in rows}
    for name in missing_sheet:
        if name not in known:
            skipped.append({'name': name, 'cohort': '', 'reason': "listed in 'כתובות מייל חסרות'"})
    skipped_path = os.path.join(ROSTER_DIR, 'skipped.csv')
    with open(skipped_path, 'w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=['name', 'cohort', 'reason'])
        writer.writeheader()
        writer.writerows(skipped)

    from_linkedin = sum(1 for r in rows if r['source'] == 'linkedin')
    print('importable people      : %d  -> %s' % (len(rows), review_path))
    print('  english from linkedin: %d' % from_linkedin)
    print('  transliterated       : %d  (needs_review=yes)' % (len(rows) - from_linkedin))
    print('skipped, cannot log in : %d  -> %s' % (len(skipped), skipped_path))
    print('\nreview %s, fix the full_name_en column, then run import_alumni.mjs'
          % os.path.basename(review_path))


if __name__ == '__main__':
    main()
